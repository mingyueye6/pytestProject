import json
import re
import pytest
import requests
from openpyxl import load_workbook


class operationExcel:
    def get_values(self, filename, sheetname=None):
        '''
        获取所有值
        :param filename: 文件路径
        :param sheetname: 工作表名称
        :return: 迭代对象(每行的值)
        '''
        wb = load_workbook(filename)
        if sheetname:
            ws = wb[sheetname]
        else:
            ws = wb[wb.sheetnames[0]]
        return ws.values

    def get_value(self, filename, coordinate, sheetname=None):
        '''
        :param filename: 文件路径
        :param coordinate: 单元格，例如[坐标(A2] 或者 [行，列]
        :param sheetname:  工作表名称
        :return: 单元格的值
        '''
        wb = load_workbook(filename)
        if sheetname:
            ws = wb[sheetname]
        else:
            ws = wb[wb.sheetnames[0]]
        if len(coordinate) == 1:
            value = ws[coordinate[0]]
        elif len(coordinate) == 2:
            value = ws.cell(row=coordinate[0], column=coordinate[1])
        else:
            value = ""
        return value

    def get_cells(self, filename, sheetname=None):
        '''
        :param filename: 文件路径
        :param sheetname: 工作表名称
        :return: 所有单元格列表
        '''
        wb = load_workbook(filename)
        if sheetname:
            ws = wb[sheetname]
        else:
            ws = wb[wb.sheetnames[0]]
        file_list = []
        for row in ws.rows:
            ls = []
            for cell in row:
                ls.append(cell.coordinate)
            file_list.append(ls)
        return file_list

    def update_value(self, filename, value, sheetname=None):
        '''
        :param filename: 文件路径
        :param value: 需要修改的单元格，例如：[坐标(例A2),值] 或 [行,列,值]
        :param sheetname: 工作表名称
        :return: code: 0修改成功，1修改失败
        '''
        wb = load_workbook(filename)
        if sheetname:
            ws = wb[sheetname]
        else:
            ws = wb[wb.sheetnames[0]]
        status = {"code": 0, "message": "修改成功"}
        if len(value) == 2:
            ws[value[0]] = value[1]
        elif len(value) == 3:
            ws.cell(row=value[0], column=value[1], value=value[2])
        else:
            status["code"] = 1
            status["message"] = "参数传递错误，%s" % value
            return status
        wb.save(filename)
        return status

    def update_values(self, filename, values, sheetname=None):
        '''
        :param filename: 文件路径
        :param values: 需要修改的单元格列表，例如：[[坐标(例A2),值],[行,列,值]]
        :param sheetname: 工作表名称
        :return: code: 0修改成功，1修改失败
        '''
        wb = load_workbook(filename)
        if sheetname:
            ws = wb[sheetname]
        else:
            ws = wb[wb.sheetnames[0]]
        status = {"code": 0, "message": ""}
        for value in values:
            if len(value) == 2:
                ws[value[0]] = value[1]
            elif len(value) == 3:
                ws.cell(row=value[0], column=value[1], value=value[2])
            else:
                status["code"] = 1
                status["message"] = "参数传递错误，%s" % value
                return status
        wb.save(filename)
        return status


if __name__ == '__main__':
    excel = operationExcel()
    print(help(excel.get_values))
    # excel.get_values("../data/users.xlsx")
    # excel.get_value("../data/users.xlsx", coordinate=["C2"])
    # excel.get_cells("../data/users.xlsx")
    # excel.update_value("../data/users.xlsx", value=[2,3,"执行通过"])
    # excel.update_values("../data/users.xlsx", values=[[3, 3, "执行失败"],[4,3, "执行失败"]])
