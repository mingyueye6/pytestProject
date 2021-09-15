import datetime
import os

from openpyxl import load_workbook
import xlrd
from xlrd import xldate_as_tuple
from xlutils.copy import copy


class ReadExcel:
    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def get_all_values(self):
        wb = load_workbook(self.filename)
        if self.sheetname:
            ws = wb[self.sheetname]
        else:
            ws = wb[wb.sheetnames[0]]
        data = []
        # for row in range(2, ws.max_row + 1):
        #     row_data = []
        #     for col in range(1, ws.max_column + 1):
        #         value = ws.cell(row, col).value
        #         c_type = ws.cell(row, col).data_type
        #         if c_type == "d":
        #             value = value.strftime('%Y-%m-%d %H:%M')
        #         row_data.append(value)
        #     data.append(row_data)
        for row_index, row_value in enumerate(ws.values):
            ls = [row_index]
            for value in row_value:
                ls.append(value)
            data.append(ls)
        data = data[1:]
        return data

    def get_cell_value(self, row, col):
        wb = load_workbook(self.filename)
        if self.sheetname:
            ws = wb[self.sheetname]
        else:
            ws = wb[wb.sheetnames[0]]
        value = ws.cell(row, col).value
        # c_type = ws.cell(row, col).data_type
        # if c_type == "d":
        #     value = value.strftime('%Y-%m-%d %H:%M')
        return value

    def write_value(self, row, col, value):
        wb = load_workbook(self.filename)
        if self.sheetname:
            ws = wb[self.sheetname]
        else:
            ws = wb[wb.sheetnames[0]]
        ws.cell(row, col, value)
        wb.save(self.filename)


class OperationExcel:
    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def get_all_values(self):
        data = xlrd.open_workbook(self.filename)
        if self.sheetname:
            sheet = data.sheet_by_name(self.sheetname)
        else:
            sheet = data.sheet_by_index(0)
        values = []
        for i in range(1, sheet.nrows):
            row_value = []
            for j in range(sheet.ncols):
                value = sheet.cell(i, j).value
                # c_type = sheet.cell(i, j).ctype
                # if c_type == 2 and value % 1 == 0:
                #     value = int(value)
                # elif c_type == 3:
                #     date = datetime.datetime(*xldate_as_tuple(value, 0))
                #     value = date.strftime('%Y/%d/%m %H:%M:%S')
                # elif c_type == 4:
                #     value = True if value == 1 else False
                row_value.append(value)
            values.append(row_value)
        return values

    def get_cell_value(self, row, col):
        data = xlrd.open_workbook(self.filename)
        if self.sheetname:
            sheet = data.sheet_by_name(self.sheetname)
        else:
            sheet = data.sheet_by_index(0)
        value = sheet.cell(row, col).value
        # c_type = sheet.cell(row, col).ctype
        # if c_type == 2 and value % 1 == 0:
        #     value = int(value)
        # elif c_type == 3:
        #     date = datetime.datetime(*xldate_as_tuple(value, 0))
        #     value = date.strftime('%Y/%d/%m %H:%M:%S')
        # elif c_type == 4:
        #     value = True if value == 1 else False
        return value

    def write_value(self, row, col, value):
        read_data = xlrd.open_workbook(self.filename)
        write_data = copy(read_data)
        sheet_data = write_data.get_sheet(0)
        sheet_data.write(row, col, value)
        write_data.save(self.filename)


if __name__ == '__main__':
    excel = ReadExcel("../data/test_case.xlsx", "case")
    print(excel.get_all_values())
    print(excel.get_cell_value(2, 13))
    # excel.write_value(2, 13, "pass")
    # print(excel.get_cell_value(2, 13))

    # opers = OperationExcel("../data/test_case_copy.xlsx", "case")
    # print(opers.get_all_values())
    # print(opers.get_cell_value(1, 1))
    # print(opers.write_value(1, 2, "pass"))
